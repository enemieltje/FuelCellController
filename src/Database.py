import io
import sqlite3
import logging
from openpyxl import load_workbook
from openpyxl.chart import Series, Reference, ScatterChart
from openpyxl.drawing.line import LineProperties
import pandas as pd
import threading
from datetime import datetime
from queue import Queue
from pathlib import Path
from enum import IntEnum

logger = logging.getLogger(__name__)


class SENSOR_ID(IntEnum):
    FUELCELL_POWER = 1
    FUELCELL_VOLTAGE = 2
    FUELCELL_CURRENT = 3
    BATTERY_POWER = 4
    BATTERY_VOLTAGE = 5
    BATTERY_CURRENT = 6
    LOAD_POWER = 7
    LOAD_VOLTAGE = 8
    LOAD_CURRENT = 9
    BATTERY_SOC = 10
    PRESSURE = 11
    THRUST = 12
    THROTTLE = 13


SENSOR_COLORS = {
    "FUELCELL_POWER": "1f77b4",
    "FUELCELL_VOLTAGE": "aec7e8",
    "FUELCELL_CURRENT": "ff7f0e",
    "BATTERY_POWER": "2ca02c",
    "BATTERY_VOLTAGE": "98df8a",
    "BATTERY_CURRENT": "d62728",
    "LOAD_POWER": "9467bd",
    "LOAD_VOLTAGE": "c5b0d5",
    "LOAD_CURRENT": "8c564b",
    "BATTERY_SOC": "e377c2",
    "PRESSURE": "7f7f7f",
    "THRUST": "bcbd22",
    "THROTTLE": "17becf",
}


class Database:

    def start():
        logger.info("Starting Database")
        Database.queue = Queue()
        Database.running = True
        Database.current_run = None

        Database.thread = threading.Thread(target=Database.worker, daemon=True)
        Database.thread.start()

    def stop():
        Database.stop_run()
        Database.running = False
        Database.queue.put({
            "action": "stop",
        })
        Database.thread.join()

    def worker():
        Database.start_db()
        # Database.current_run = Database.create_run()
        # logger.debug(f"Started Run:\n{Database.current_run}")

        while Database.running:
            request = Database.queue.get()

            action = request["action"]

            if action == "insert":
                Database._insert(request)

            elif action == "get_latest":
                Database._get_latest(request)

            elif action == "get_csv":
                Database._get_csv(request)

            elif action == "get_current_run":
                Database._get_current_run(request)

            elif action == "get_runs":
                Database._get_runs(request)

            elif action == "get_run_name":
                Database._get_run_name(request)

            elif action == "start_run":
                Database._start_run(request)

            elif action == "update_run":
                Database._update_run(request)

            elif action == "stop_run":
                Database._stop_run(request)

            elif action == "export_run_csv":
                Database._export_run_csv(request)

            elif action == "export_run_excel":
                Database._export_run_excel(request)

            elif action == "stop":
                Database.stop_db()

    def start_db(filename="data/sensors.db"):
        logger.debug("Connecting to Database...")
        Database.conn = sqlite3.connect(filename)
        Database.cursor = Database.conn.cursor()

        sql = Database.load_sql("create_run_table")
        Database.cursor.execute(sql)

        sql = Database.load_sql("create_sample_table")
        Database.cursor.execute(sql)

        Database.conn.commit()
        logger.debug("Database Connected!")

    def stop_db():
        Database.conn.close()

    def create_run():
        sql = Database.load_sql("insert_run")
        started_at = datetime.now().isoformat()

        Database.cursor.execute(sql, (
            f"Run {started_at}",
            started_at,
        ))
        Database.conn.commit()

        return Database.cursor.lastrowid

    def row_to_dict(row):
        if row is None:
            return None

        return {
            description[0]: value
            for description, value in zip(Database.cursor.description, row)
        }

    def request(action, **kwargs):
        response_queue = Queue()
        kwargs["action"] = action
        kwargs["response_queue"] = response_queue
        Database.queue.put(kwargs)
        return response_queue.get()

    def load_sql(name):
        return Path(f"src/sql/{name}.sql").read_text()

    def get_all_sensors():
        logger.debug("Getting All Sensors")
        sensor_data = {sensor.name: Database.get_latest(
            sensor.value) for sensor in SENSOR_ID}
        logger.debug(sensor_data)
        return sensor_data

    def insert(sensor_id, value):
        if Database.current_run is None:
            return

        # logger.debug(f"Inserting sensor {sensor_id}")
        Database.queue.put({
            "action": "insert",
            "run_id": Database.current_run,
            "sensor_id": sensor_id,
            "value": value,
        })

    def _insert(request):
        sql = Database.load_sql("insert_sample")

        Database.cursor.execute(sql, (
            request["run_id"],
            datetime.now().isoformat(),
            request["sensor_id"],
            request["value"]
        ))
        Database.conn.commit()

    def get_latest(sensor_id):
        logger.debug(f"Getting latest sensor {sensor_id}")
        response_queue = Queue()

        Database.queue.put({
            "action": "get_latest",
            "sensor_id": sensor_id,
            "response_queue": response_queue
        })

        return response_queue.get()  # blocks until result arrives

    def _get_latest(request):
        sql = Database.load_sql("get_latest")

        Database.cursor.execute(sql, (request["sensor_id"],))
        result = Database.cursor.fetchone()
        if result is None:
            result = (0, 0, "", 0, 0)
        logger.debug(result)

        if request.get("response_queue"):
            request["response_queue"].put(result[4])

    def get_current_run():
        return Database.request("get_current_run")

    def _get_current_run(request):
        run = None

        if Database.current_run is not None:
            Database.cursor.execute(
                "SELECT * FROM test_runs WHERE id = ?",
                (Database.current_run,)
            )
            run = Database.row_to_dict(Database.cursor.fetchone())

        request["response_queue"].put(run)

    def get_runs():
        # logger.debug("Getting Runs")
        return Database.request("get_runs")

    def _get_runs(request):
        sql = Database.load_sql("get_runs")

        Database.cursor.execute(sql)
        rows = Database.cursor.fetchall()
        columns = [description[0]
                   for description in Database.cursor.description]
        runs = [
            {column: value for column, value in zip(columns, row)}
            for row in rows
        ]

        request["response_queue"].put(runs)

    def start_run(name=None, notes=None):
        return Database.request("start_run", name=name, notes=notes)

    def _start_run(request):
        if Database.current_run is not None:
            Database._stop_run({
                "run_id": Database.current_run,
                "response_queue": None,
            })

        Database.current_run = Database.create_run()
        logger.info(f"Started New Run: {Database.current_run}")

        if request.get("name") is not None or request.get("notes") is not None:
            Database._update_run({
                "run_id": Database.current_run,
                "name": request.get("name"),
                "notes": request.get("notes"),
                "response_queue": None,
            })

        Database._get_current_run(request)

    def update_run(run_id, name=None, notes=None):
        return Database.request("update_run", run_id=run_id, name=name, notes=notes)

    def _update_run(request):
        sql = Database.load_sql("update_run")

        Database.cursor.execute(sql, (
            request.get("name"),
            request.get("notes"),
            request["run_id"],
        ))
        Database.conn.commit()

        if request.get("response_queue"):
            Database.cursor.execute(
                "SELECT * FROM test_runs WHERE id = ?",
                (request["run_id"],)
            )
            request["response_queue"].put(
                Database.row_to_dict(Database.cursor.fetchone()))

    def stop_run(run_id=None):
        return Database.request("stop_run", run_id=run_id)

    def _stop_run(request):
        run_id = request.get("run_id") or Database.current_run

        if run_id is None:
            if request.get("response_queue"):
                request["response_queue"].put(None)
            return

        logger.info(f"Stopped Run: {run_id}")
        sql = Database.load_sql("stop_run")
        Database.cursor.execute(sql, (
            datetime.now().isoformat(),
            run_id,
        ))
        Database.conn.commit()

        if run_id == Database.current_run:
            Database.current_run = None

        if request.get("response_queue"):
            Database.cursor.execute(
                "SELECT * FROM test_runs WHERE id = ?",
                (run_id,)
            )
            request["response_queue"].put(
                Database.row_to_dict(Database.cursor.fetchone()))

    def export_run_csv(run_id):
        Database.export_run_excel(run_id)
        return Database.request("export_run_csv", run_id=run_id)

    def _export_run_csv(request):
        csv_text = Database._get_csv(request)
        request["response_queue"].put(csv_text)

    def get_csv():
        logger.debug(f"Requesting csv ({Database.current_run})")
        Database.queue.put({
            "action": "get_csv",
            "run_id": Database.current_run,
        })

    def _get_df(request):
        sql = Database.load_sql("get_run_samples")

        df = pd.read_sql_query(sql, Database.conn, params=(request["run_id"],))
        df["timestamp"] = pd.to_datetime(df["timestamp"])

        pivot_df = df.pivot(
            index="timestamp",
            columns="sensor_id",
            values="value")

        pivot_df = pivot_df.ffill()
        pivot_df = pivot_df.rename(
            columns={sensor.value: sensor.name for sensor in SENSOR_ID})
        return pivot_df

    def _get_csv(request):
        logger.debug(f"Getting csv ({request["run_id"]})")
        df = Database._get_df(request)
        return df.to_csv(index=True)

    def export_run_excel(run_id):
        return Database.request("export_run_excel", run_id=run_id)

    def _export_run_excel(request):
        logger.debug(f"Getting excel ({request["run_id"]})")
        df = Database._get_df(request)

        # Write to in-memory buffer
        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=True, sheet_name="run_data")

        buffer.seek(0)

        # -----------------------------------
        # Create Excel chart
        # -----------------------------------

        wb = load_workbook(buffer)
        ws = wb.active

        Database.add_sensor_chart(
            ws,
            title="Power vs Throttle",
            primary_axis_title="Power (W)",
            primary_sensors=[
                "FUELCELL_POWER",
                "BATTERY_POWER",
                "LOAD_POWER"
            ],
            secondary_axis_title="Throttle",
            secondary_sensors=[
                "THROTTLE"
            ],
            chart_position="O4"
        )

        Database.add_sensor_chart(
            ws,
            title="Fuelcell Power",
            primary_axis_title="Power (W)",
            primary_sensors=[
                "FUELCELL_POWER",
            ],
            secondary_axis_title="Voltage (v), Current (A)",
            secondary_sensors=[
                "FUELCELL_VOLTAGE",
                "FUELCELL_CURRENT",
            ],
            chart_position="O34"
        )

        final_buffer = io.BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        request["response_queue"].put(final_buffer.read())

    def add_sensor_chart(
        ws,
        title: str,
        primary_sensors: list[str],
        secondary_sensors: list[str],
        primary_axis_title="Primary Axis",
        secondary_axis_title="Secondary Axis",
        x_col: int = 1,
        start_row: int = 2,
        chart_position: str = "O4",
    ):
        """
        Adds a scatter chart to an Excel worksheet.

        Parameters:
        - ws: openpyxl worksheet
        - title: chart title
        - primary_sensors: list of column headers on primary Y-axis
        - secondary_sensors: list of column headers on secondary Y-axis
        """

        chart = ScatterChart()
        chart.title = title
        chart.scatterStyle = "lineMarker"

        chart.x_axis.title = "Timestamp"
        chart.x_axis.number_format = "hh:mm:ss"
        chart.x_axis.delete = False
        chart.y_axis.title = primary_axis_title
        chart.y_axis.delete = False
        chart.legend.position = "r"
        chart.width = 48
        chart.height = 14

        # Secondary axis setup
        secondary_chart = ScatterChart()
        secondary_chart.scatterStyle = "lineMarker"

        secondary_chart.y_axis.axId = 200
        secondary_chart.y_axis.title = secondary_axis_title
        secondary_chart.y_axis.crosses = "max"
        secondary_chart.y_axis.delete = False

        # X-axis values (timestamp column)
        xvalues = Reference(
            ws,
            min_col=x_col,
            min_row=start_row,
            max_row=ws.max_row
        )

        def add_series(sensor_name, target_chart):
            # find column by header row
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == sensor_name:
                    values = Reference(
                        ws,
                        min_col=col,
                        min_row=start_row,
                        max_row=ws.max_row
                    )

                    series = Series(values, xvalues, title=sensor_name)

                    color = SENSOR_COLORS[sensor_name]

                    if color:
                        series.graphicalProperties.line = LineProperties(
                            solidFill=color
                        )

                    target_chart.series.append(series)
                    return

            raise ValueError(f"Sensor '{sensor_name}' not found in worksheet")

        # Add primary sensors
        for sensor in primary_sensors:
            add_series(sensor, chart)

        # Add secondary sensors
        for sensor in secondary_sensors:
            add_series(sensor, secondary_chart)

        # Combine charts
        chart += secondary_chart

        # Add to worksheet
        ws.add_chart(chart, chart_position)

    def get_run_name(run_id=None):
        return Database.request("get_run_name", run_id=run_id)

    def _get_run_name(request):
        logger.debug(f"Getting run name ({request["run_id"]})")
        run_id = request.get("run_id") or Database.current_run
        sql = Database.load_sql("get_run_name")

        Database.cursor.execute(sql, (
            run_id,
        ))
        result = Database.cursor.fetchone()
        if result is None:
            result = (0, 0)

        request["response_queue"].put(result[1])
